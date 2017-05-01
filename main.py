import os
import time

import openpyxl

from util import util
from crawler.scraping import scraping

welcome = '''
//////////////////////////////////////
//         知网论文爬取工具           //
//             2017-5-1             //
//          @author firejq          //
//////////////////////////////////////
            程序正在初始化……
'''
print(welcome)

driver = util.getDriver('phantomjs')
url = 'http://kns.cnki.net/kns/brief/result.aspx?dbprefix=scdb'

print('请输入论文发表时间范围（若无时间范围限制，直接回车即可）：')
begin_time = util.get_time_input('开始时间（1979-01-01 ~ 2017-12-31）：')
end_time = util.get_time_input('截止时间（1979-01-01 ~ 2017-12-31）：')

authors = util.get_name_list()

path = os.getcwd() + os.sep + 'result'
filename = 'result.xlsx'
out_path = os.path.join(path, filename)

if not os.path.exists(out_path):
    # 程序未运行过的情况
    if not os.path.exists(path):
        os.mkdir(path)
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(index=1, title='tmp')
    # 初始化
    for author in authors:
        tmp = [author, 'unsolved']
        line = [l for l in tmp]
        ws.append(line)
    wb.save(out_path)
    for author in authors:
        driver.get(url=url)
        time.sleep(1)
        scraping(driver=driver, author_name=author, author_company='深圳大学', begin_time=begin_time, end_time=end_time)
else:
    # 程序已经运行过
    wb = openpyxl.load_workbook(out_path)
    sheet_names = wb.get_sheet_names()
    if 'tmp' in sheet_names:
        # 程序上次运行到一半意外退出了
        ws = wb.get_sheet_by_name('tmp')
        for i in range(1, ws.max_row + 1):
            if ws['B' + str(i)].value == 'solved':
                print('【' + ws['A' + str(i)].value + '】的信息已经抓取过，进行下一个作者的信息抓取')
                continue
            if ws['B' + str(i)].value == 'unsolved':
                driver.get(url=url)
                time.sleep(1)
                scraping(driver=driver, author_name=ws['A' + str(i)].value, author_company='深圳大学',
                         begin_time=begin_time, end_time=end_time)
    else:
        # 程序上次全程运行完毕后才正确退出
        ws = wb.create_sheet(index=1, title='tmp')
        # 初始化
        for author in authors:
            tmp = [author, 'unsolved']
            line = [l for l in tmp]
            ws.append(line)
        wb.save(out_path)
        for author in authors:
            driver.get(url=url)
            time.sleep(1)
            scraping(driver=driver, author_name=author, author_company='深圳大学', begin_time=begin_time, end_time=end_time)

util.delete_sheet_tmp()
driver.quit()
print('名单上所有作者的论文信息抓取完毕')



# TODO 去重
# TODO 提高性能--多线程 & 多进程 & 异步IO
