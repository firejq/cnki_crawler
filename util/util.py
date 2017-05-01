import os
import random
import re

import openpyxl
from selenium import webdriver
from selenium.webdriver import DesiredCapabilities


def get_name_list():
    # 直接从内存中获取名单列表
    authors = [
        "陈思平", "陈昕", "汪天富", "谭力海", "彭珏", "但果", "叶继伦", "覃正笛",
        "张旭", "张会生", "钱建庭", "丁惠君", "刁现芬", "沈圆圆", "周永进", "孔湉湉",
        "陆敏华", "张新宇", "孙怡雯", "李乔亮", "齐素文", "徐海华", "倪东", "刘维湘",
        "李抱朴", "黄炳升", "徐敏", "雷柏英", "胡亚欣", "何前军", "郑介志", "常春起",
        "陈雯雯", "罗永祥", "黄鹏", "林静", "王倪传", "刘立", "张治国", "董磊"
    ]

    # 从name-list.txt获取名单
    # name_list_file = os.getcwd() + os.sep + 'docs' + os.sep + 'name-list.txt'
    # authors = open(name_list_file, 'r', encoding='utf-8').readlines()
    # # 去除每行行尾的‘\n’
    # for i in range(0, len(authors)):
    #     authors[i] = authors[i].strip('\n')
    return authors

def getDriver(browser='chrome'):
    UserAgents = [
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3072.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:46.0) Gecko/20100101 Firefox/46.0',
        'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.87 Safari/537.36 OPR/37.0.2178.32',
        'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/534.57.2 (KHTML, like Gecko) Version/5.1.7 Safari/534.57.2',
        'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/45.0.2454.101 Safari/537.36',
        'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/45.0.2454.101 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2486.0 Safari/537.36 Edge/13.10586',
        'Mozilla/5.0 (Windows NT 10.0; WOW64; Trident/7.0; rv:11.0) like Gecko',
        'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/47.0.2526.106 BIDUBrowser/8.3 Safari/537.36',
        'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Maxthon/4.9.2.1000 Chrome/39.0.2146.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 61; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/47.0.2526.80 Safari/537.36 Core/1.47.277.400 QQBrowser/9.4.7658.400',
        'Mozilla/5.0 (Linux; Android 5.0; SM-N9100 Build/LRX21V) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/37.0.0.0 Mobile Safari/537.36 MicroMessenger/6.0.2.56_r958800.520 NetType/WIFI',
        'Mozilla/5.0 (iPhone; CPU iPhone OS 7_1_2 like Mac OS X) AppleWebKit/537.51.2 (KHTML, like Gecko) Mobile/11D257 QQ/5.2.1.302 NetType/WIFI Mem/28'
    ]
    if browser == 'chrome':
        options = webdriver.ChromeOptions()
        options.add_argument('user-agent=' + random.choice(UserAgents))
        executable_path = os.getcwd() + os.sep + 'chromedriver.exe'
        driver = webdriver.Chrome(executable_path=executable_path,
                                  chrome_options=options)
    elif browser == 'phantomjs':
        dcap = dict(DesiredCapabilities.PHANTOMJS)
        dcap["phantomjs.page.settings.userAgent"] = (
            random.choice(UserAgents)
        )
        executable_path = os.getcwd() + os.sep + 'phantomjs.exe'
        driver = webdriver.PhantomJS(
            executable_path=executable_path,
            desired_capabilities=dcap)
    return driver


# def is_done(author_name):
#     '''
#     后期使用数据库查重，可代替此函数
#     :param author_name:
#     :return:
#     '''
#     path = os.getcwd() + os.sep + 'result'
#     filename = 'result.xlsx'
#     out_path = os.path.join(path, filename)
#     if os.path.exists(out_path):
#         wb = openpyxl.load_workbook(out_path)
#         sheets_name = wb.get_sheet_names()
#         if author_name in sheets_name:
#             return True
#         else:
#             return False
#     else:
#         return False


def write_to_excel(res):
    path = os.getcwd() + os.sep + 'result'
    filename = 'result.xlsx'
    out_path = os.path.join(path, filename)

    wb = openpyxl.load_workbook(out_path)
    sheet_names = wb.get_sheet_names()
    if 'result' in sheet_names:
        ws = wb.get_sheet_by_name('result')
        # 检查该条记录是否已经存在
        max_row = ws.max_row
        for index in range(2, max_row + 1):
            if res[0] == ws['A' + str(index)].value:
                print('~~~~~~~~该条记录在结果集中已存在，进行下一条~~~~~~~~~')
                return
    else:
        ws = wb.create_sheet(index=0, title='result')
        titles = ['题名', '作者', '来源', '发表时间', '数据库']
        line = [title for title in titles]
        ws.append(line)

    line = [r for r in res]
    ws.append(line)
    wb.save(out_path)
    # print('成功添加一条【' + author_name + '】的记录')


def change_tmp_status(author_name):
    file_path = os.getcwd() + os.sep + 'result' + os.sep + 'result.xlsx'
    wb = openpyxl.load_workbook(file_path)
    ws = wb.get_sheet_by_name('tmp')

    for i in range(1, ws.max_row+1):
        # print(ws['A'+str(i)])
        # continue
        if ws['A'+str(i)].value == author_name:
            ws['B'+str(i)].value ='solved'
            wb.save(file_path)
            break


def delete_sheet_tmp():
    file_path = os.getcwd() + os.sep + 'result' + os.sep + 'result.xlsx'
    wb = openpyxl.load_workbook(file_path)
    wb.remove_sheet(wb.get_sheet_by_name('tmp'))
    wb.save(file_path)


def get_time_input(message):
    while(True):
        vaule = input(message)
        if vaule == '':
            return vaule
        # elif re.match(r'^\d{4}(-\d\d){2}$', vaule):
        elif re.match(r'^((19(79|[89][0-9]))|(20(0[0-9]|1[0-7])))-(0[1-9]|1[0-2])-(0[1-9]|1[0-9]|2[0-9]|3[0-1])$', vaule):
            return vaule
        else:
            print('输入的日期不合法，请重新输入！')

